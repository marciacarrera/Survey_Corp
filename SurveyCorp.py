
from openpyxl import load_workbook

def contar_estados_excel():
    # Cargar el archivo Excel
    workbook = load_workbook('Survey Corp_Caso_Cliente_.xlsx')

    # Obtener la primera hoja del libro
    sheet = workbook.active

    # Buscar los índices de las columnas 'Estado' (columna E) y 'Tipo' (columna B)
    estado_column_index = 5  # Índice de la columna E
    tipo_column_index = 2  # Índice de la columna B

    # Crear diccionarios para almacenar los conteos
    conteo_int = {'pending': 0, 'done': 0}
    conteo_ext = {'pending': 0, 'done': 0}

    # Recorrer las filas y realizar el conteo
    for row in range(2, sheet.max_row + 1):
        estado = sheet.cell(row=row, column=estado_column_index).value
        tipo = sheet.cell(row=row, column=tipo_column_index).value

        if tipo.lower().startswith('int'):
            if estado == 'pending':
                conteo_int['pending'] += 1
            elif estado == 'done':
                conteo_int['done'] += 1
        elif tipo.lower().startswith('ext'):
            if estado == 'pending':
                conteo_ext['pending'] += 1
            elif estado == 'done':
                conteo_ext['done'] += 1

    # Imprimir los resultados
    print("Conteo para 'Int.'")
    print("Número de 'pending':", conteo_int['pending'])
    print("Número de 'done':", conteo_int['done'])

    print("\nConteo para 'Ext.'")
    print("Número de 'pending':", conteo_ext['pending'])
    print("Número de 'done':", conteo_ext['done'])

# Llamar a la función contar_estados_excel
contar_estados_excel()
